"""Tests for the calendar-year overview (modules.annual) and the annual
Excel/PDF reports."""

from datetime import date

from openpyxl import load_workbook

from modules import annual
from modules.db_handler.database import Database
from modules.db_handler.repositories import (
    FixedCostRepository,
    IncomeRepository,
    VariableExpenseRepository,
    VariableIncomeRepository,
)
from modules.file_handler import excel_io, pdf_report
from modules.models import FixedCost, IncomeSource, VariableExpense, VariableIncome

TODAY = date(2026, 7, 15)


def _repos(tmp_path):
    db = Database(tmp_path / "test.db")
    return (db, IncomeRepository(db), FixedCostRepository(db),
            VariableExpenseRepository(db), VariableIncomeRepository(db))


def _seed(income, fixed, exp, var_inc):
    income.add(IncomeSource("Job", 200_000, "vollzeit"))          # every month
    fixed.add(FixedCost("Miete", 80_000, "Wohnen"))               # open-ended
    exp.add(VariableExpense("2025-03-10", 30_000, "Lebensmittel"))
    exp.add(VariableExpense("2025-08-05", 20_000, "Auto & Tanken"))
    exp.add(VariableExpense("2026-02-14", 25_000, "Lebensmittel"))
    exp.add(VariableExpense("2026-05-01", 10_000, "Freizeit & Unterhaltung"))
    var_inc.add(VariableIncome("2026-03-31", 15_000, "Rückzahlung"))


def test_build_totals_and_categories(tmp_path):
    db, income, fixed, exp, var_inc = _repos(tmp_path)
    _seed(income, fixed, exp, var_inc)

    ov = annual.build(income, fixed, exp, var_inc, 2025, today=TODAY)
    assert ov.months_covered == 12                       # past year = full year
    assert ov.income_cents == 12 * 200_000
    assert ov.fixed_cents == 12 * 80_000
    assert ov.variable_cents == 50_000
    assert ov.remaining_cents == ov.income_cents - ov.fixed_cents - 50_000
    assert ov.by_category == {"Lebensmittel": 30_000, "Auto & Tanken": 20_000}
    assert ov.has_data
    db.close()


def test_current_year_covers_only_elapsed_months(tmp_path):
    db, income, fixed, exp, var_inc = _repos(tmp_path)
    _seed(income, fixed, exp, var_inc)

    ov = annual.build(income, fixed, exp, var_inc, 2026, today=TODAY)
    assert ov.months_covered == 7                        # Januar..Juli
    assert ov.income_cents == 7 * 200_000 + 15_000       # incl. one-off in March
    assert ov.variable_cents == 35_000
    assert len(ov.points) == 7
    # A future year has no coverage at all.
    assert annual.months_covered(2027, TODAY) == 0
    db.close()


def test_savings_rate_and_no_income_case(tmp_path):
    db, income, fixed, exp, var_inc = _repos(tmp_path)
    ov = annual.build(income, fixed, exp, var_inc, 2025, today=TODAY)
    assert ov.savings_rate is None                       # no income at all
    _seed(income, fixed, exp, var_inc)
    ov = annual.build(income, fixed, exp, var_inc, 2025, today=TODAY)
    expected = ov.remaining_cents / ov.income_cents
    assert abs(ov.savings_rate - expected) < 1e-9
    db.close()


def test_category_matrix_lines_up(tmp_path):
    db, income, fixed, exp, var_inc = _repos(tmp_path)
    _seed(income, fixed, exp, var_inc)
    ov = annual.build(income, fixed, exp, var_inc, 2025, today=TODAY)
    categories, per_month = annual.category_matrix(ov)
    assert categories == ["Lebensmittel", "Auto & Tanken"]
    assert len(per_month["Lebensmittel"]) == 12
    assert per_month["Lebensmittel"][2] == 30_000        # März
    assert per_month["Auto & Tanken"][7] == 20_000       # August
    assert sum(per_month["Lebensmittel"]) == 30_000
    db.close()


def test_available_years_span_oldest_booking(tmp_path):
    db, income, fixed, exp, var_inc = _repos(tmp_path)
    assert annual.available_years(exp, var_inc, today=TODAY) == [2026]
    _seed(income, fixed, exp, var_inc)
    assert annual.available_years(exp, var_inc, today=TODAY) == [2026, 2025]
    exp.add(VariableExpense("2019-01-05", 1_000, "Sonstiges"))
    years = annual.available_years(exp, var_inc, today=TODAY)
    assert years[0] == 2026 and years[-1] == 2019
    db.close()


def test_year_excel_report(tmp_path):
    db, income, fixed, exp, var_inc = _repos(tmp_path)
    _seed(income, fixed, exp, var_inc)
    ov = annual.build(income, fixed, exp, var_inc, 2026, today=TODAY)
    prev = annual.build(income, fixed, exp, var_inc, 2025, today=TODAY)

    out = excel_io.export_year_workbook(
        tmp_path / "jahr.xlsx", overview=ov, prev_overview=prev)
    assert out.exists()
    wb = load_workbook(out)
    assert {"Jahresübersicht", "Monate", "Monatsmatrix"} <= set(wb.sheetnames)
    # The comparison header names both years.
    heads = [c.value for c in wb["Jahresübersicht"][4]]
    assert "2026" in heads and "2025" in heads
    db.close()


def test_year_pdf_report(tmp_path):
    db, income, fixed, exp, var_inc = _repos(tmp_path)
    _seed(income, fixed, exp, var_inc)
    ov = annual.build(income, fixed, exp, var_inc, 2026, today=TODAY)
    prev = annual.build(income, fixed, exp, var_inc, 2025, today=TODAY)

    out = pdf_report.generate_year_report(
        tmp_path / "jahr.pdf", overview=ov, prev_overview=prev)
    assert out.exists() and out.stat().st_size > 1_000
    assert out.read_bytes()[:5] == b"%PDF-"
    # Without a previous year the simple table variant is used — must not crash.
    out2 = pdf_report.generate_year_report(tmp_path / "solo.pdf", overview=ov)
    assert out2.exists()
    db.close()
