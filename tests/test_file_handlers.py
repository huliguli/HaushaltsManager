"""End-to-end tests for the Excel/PDF import-export layer."""

from openpyxl import load_workbook

from modules.file_handler import excel_io, pdf_import, pdf_report
from modules.models import BudgetOverview, Credit, FixedCost, VariableExpense


def _sample():
    overview = BudgetOverview(
        income_cents=230_000, fixed_cents=95_000, variable_cents=68_000,
        credits_cents=7_500,
        expenses_by_category={"Lebensmittel": 30_000, "Auto & Tanken": 20_000},
    )
    fixed = [
        FixedCost("Miete", 50_000, "Wohnen"),
        FixedCost("Ratenkredit", 7_500, "Kredit", end_date="2026-12-14"),
    ]
    expenses = [
        VariableExpense("2026-06-05", 3_000, "Lebensmittel", "Supermarkt"),
        VariableExpense("2026-06-10", 5_000, "Auto & Tanken", "=cmd|calc"),  # injection attempt
    ]
    credits = [Credit("Autokredit", total_cents=1_500_000, monthly_cents=25_000,
                      term_months=60, interest_rate=4.0, category="Auto")]
    return overview, fixed, expenses, credits


def test_excel_export_and_reload(tmp_path):
    overview, fixed, expenses, credits = _sample()
    out = excel_io.export_workbook(
        tmp_path / "report.xlsx", income=[], fixed_costs=fixed,
        expenses=expenses, credits=credits, overview=overview)
    assert out.exists()

    wb = load_workbook(out)
    assert {"Übersicht", "Fixkosten-Timeline", "Ausgaben", "Kredite",
            "Tilgungspläne"} <= set(wb.sheetnames)
    # The formula-injection attempt must be neutralised AND preserved: _safe_text
    # prefixes a leading apostrophe so Excel treats it as text, not a formula.
    ausgaben = wb["Ausgaben"]
    texts = [ausgaben.cell(row=r, column=3).value for r in range(2, ausgaben.max_row + 1)]
    assert "'=cmd|calc" in texts                       # neutralised, content kept
    assert all(not (t or "").startswith("=") for t in texts)  # nothing stays a formula


def test_excel_annual_saving_uses_recurring_income(tmp_path):
    # Regression: the "pro Jahr ansparbar" note must project the RECURRING
    # surplus, never the one-off-inflated month (same class as the v1.7.1 bug).
    overview = BudgetOverview(
        income_cents=500_000,            # includes a fat one-off credit this month
        recurring_income_cents=200_000,  # the real repeating income
        fixed_cents=95_000, variable_cents=68_000)
    out = excel_io.export_workbook(
        tmp_path / "annual.xlsx", income=[], fixed_costs=[], expenses=[],
        credits=[], overview=overview)
    ws = load_workbook(out)["Übersicht"]
    note = next(c.value for row in ws.iter_rows() for c in row
                if isinstance(c.value, str) and "pro Jahr ansparbar" in c.value)
    # recurring_after_all = 200000-95000-68000 = 37000 cents -> *12 = 4.440,00 €.
    assert "4.440,00" in note
    assert "40.440,00" not in note   # would be the wrong (one-off-inflated) figure


def test_excel_monthly_overview_is_month_scoped(tmp_path):
    # A monthly overview: Übersicht + Ausgaben describe the chosen month, while
    # the Monatsverlauf sheet still spans the full history (so past months can be
    # exported without losing the multi-month trend).
    month_exp = [VariableExpense("2026-06-05", 3_000, "Lebensmittel", "Juni-Kauf")]
    history = month_exp + [
        VariableExpense("2026-05-10", 5_000, "Auto & Tanken"),
        VariableExpense("2026-07-01", 9_900, "Freizeit & Unterhaltung")]
    overview = BudgetOverview(
        income_cents=200_000, recurring_income_cents=200_000,
        fixed_cents=80_000, variable_cents=3_000,
        expenses_by_category={"Lebensmittel": 3_000})
    out = excel_io.export_workbook(
        tmp_path / "monat.xlsx", income=[], fixed_costs=[], expenses=month_exp,
        history_expenses=history, credits=[], overview=overview, month_label="Juni 2026")

    wb = load_workbook(out)
    subtitle = " ".join(str(c.value) for row in wb["Übersicht"].iter_rows()
                        for c in row if c.value)
    assert "Juni 2026" in subtitle                       # month named in the overview
    aus = wb["Ausgaben"]
    descs = [aus.cell(row=r, column=3).value for r in range(1, aus.max_row + 1)]
    assert "Juni-Kauf" in descs                          # the month's expense is listed
    # Monatsverlauf spans all three history months, not just June.
    ver = wb["Monatsverlauf"]
    ym = [ver.cell(row=r, column=1).value for r in range(1, ver.max_row + 1)]
    assert len([x for x in ym if x and "/" in str(x)]) == 3


def test_excel_import_mapping_and_rows(tmp_path):
    overview, fixed, expenses, credits = _sample()
    path = excel_io.export_workbook(
        tmp_path / "r.xlsx", income=[], fixed_costs=fixed,
        expenses=expenses, credits=credits, overview=overview)
    # Re-read the Ausgaben sheet would need that sheet active; instead test the
    # column guesser directly on a typical bank-export header.
    headers = ["Buchungstag", "Verwendungszweck", "Betrag"]
    mapping = excel_io.guess_mapping(headers)
    assert mapping["date"] == 0
    assert mapping["description"] == 1
    assert mapping["amount"] == 2

    rows = [["05.06.2026", "REWE Markt", "-34,90"], ["07.06.2026", "Aral", "-52,10"]]
    items, skipped = excel_io.rows_to_expenses(rows, mapping)
    assert skipped == 0
    assert len(items) == 2
    assert items[0].amount_cents == 3_490  # absolute value of -34,90


def test_rows_to_expenses_skips_and_fallbacks():
    # The import loop's robustness branches (skip empty/garbage/zero/short rows,
    # fall back to today on a bad date, keep positive amounts) were untested.
    mapping = {"date": 0, "amount": 2, "description": 1, "category": None}
    rows = [
        ["05.06.2026", "REWE", "-34,90"],   # ok -> 3490 (abs)
        ["06.06.2026", "leer", ""],          # empty amount -> skip
        ["07.06.2026", "muell", "abc"],      # unparseable -> skip
        ["08.06.2026", "null", "0,00"],      # zero -> skip
        ["09.06.2026", "kurz"],              # too short for amount index -> skip
        ["", "kein Datum", "12,00"],         # bad date -> today fallback, still imported
        ["10.06.2026", "positiv", "55,00"],  # positive amount kept positive
    ]
    items, skipped = excel_io.rows_to_expenses(rows, mapping)
    assert skipped == 4
    assert len(items) == 3
    assert items[0].amount_cents == 3_490
    assert items[-1].amount_cents == 5_500
    # The bad-date row (12,00 € -> 1200 cents) falls back to today's ISO date.
    from modules import dates
    bad_date_item = next(it for it in items if it.amount_cents == 1_200)
    assert bad_date_item.date == dates.to_iso(dates.today())


def test_read_preview_empty_and_header_only(tmp_path):
    from openpyxl import Workbook

    empty = tmp_path / "empty.xlsx"
    Workbook().save(empty)
    headers, data, mapping = excel_io.read_preview(empty)
    assert headers == [] and data == []
    assert mapping["amount"] is None

    wb = Workbook()
    wb.active.append(["Datum", "Verwendungszweck", "Betrag"])
    header_only = tmp_path / "headeronly.xlsx"
    wb.save(header_only)
    headers, data, mapping = excel_io.read_preview(header_only)
    assert headers == ["Datum", "Verwendungszweck", "Betrag"]
    assert data == []
    assert mapping["amount"] == 2


def test_pdf_report_and_amount_extraction(tmp_path):
    overview, fixed, expenses, credits = _sample()
    out = pdf_report.generate_monthly_report(
        tmp_path / "bericht.pdf", year=2026, month=6, overview=overview,
        fixed_costs=fixed, expenses=expenses, credits=credits)
    assert out.exists()
    head = out.read_bytes()[:5]
    assert head.startswith(b"%PDF")

    # The generated report contains euro amounts; the extractor should find the
    # income figure (2.300,00 € -> 230000 cents) among them as integer cents.
    amounts = pdf_import.extract_amounts(out)
    assert len(amounts) > 0
    cents_values = [a["cents"] for a in amounts]
    assert 230_000 in cents_values
    assert all(isinstance(a["cents"], int) for a in amounts)


def test_reports_carry_the_app_version_stamp(tmp_path):
    # A report must reveal which app version produced it — that is how an
    # "old-looking" export is told apart from a caching problem (there is no
    # cache; exports are always generated live).
    from app_meta import APP_VERSION

    overview, fixed, expenses, credits = _sample()
    wb_path = excel_io.export_workbook(
        tmp_path / "stamp.xlsx", income=[], fixed_costs=fixed,
        expenses=expenses, credits=credits, overview=overview)
    subtitle = load_workbook(wb_path)["Übersicht"].cell(row=2, column=1).value
    assert f"HaushaltsManager {APP_VERSION}" in subtitle

    pdf_path = pdf_report.generate_monthly_report(
        tmp_path / "stamp.pdf", year=2026, month=6, overview=overview,
        fixed_costs=fixed, expenses=expenses, credits=credits)
    import fitz
    doc = fitz.open(pdf_path)
    text = "".join(page.get_text() for page in doc)
    doc.close()
    assert f"HaushaltsManager {APP_VERSION}" in text
