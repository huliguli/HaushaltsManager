"""Excel import and export (openpyxl).

Export writes a multi-sheet workbook (overview, fixed-cost timeline, expenses,
credits, amortisation plans). Import reads a sheet, auto-detects the column
layout of common German bank exports (DKB, Sparkasse, Commerzbank, N26, ING)
and returns an editable preview.

Security: any text cell that originates from user data is run through
``_safe_text`` to neutralise spreadsheet formula injection (a leading
=, +, - or @ is prefixed with an apostrophe so Excel treats it as text).
"""

from __future__ import annotations

from pathlib import Path

from collections import defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

from modules import dates
from modules.calculator import annuity, timeline
from modules.models import CREDIT_STATUS_LABELS, INCOME_TYPE_LABELS
from modules.money import cents_to_euros, format_eur, parse_eur

# --- Workbook styling vocabulary ------------------------------------------
# Negative amounts render red automatically (number-format section syntax).
_MONEY_FMT = '#,##0.00 €;[Red]-#,##0.00 €'
_PCT_FMT = '0.0 %;[Red]-0.0 %'
_NAVY = "1B2330"
_ACCENT = "2F6BD8"
_MUTED = "5D6877"
_GREEN = "1F9D57"
_RED = "D6453D"
_HEADER_FILL = PatternFill("solid", fgColor=_NAVY)
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_TITLE_FONT = Font(color="FFFFFF", bold=True, size=15)
_SUBTITLE_FONT = Font(color="C6D0E2", size=9.5)
_NOTE_FONT = Font(color=_MUTED, size=10)
_TOTAL_FONT = Font(color=_NAVY, bold=True)
_TOTAL_FILL = PatternFill("solid", fgColor="EEF1F6")
_ALT_FILL = PatternFill("solid", fgColor="F6F8FB")
_BANNER_FILL = PatternFill("solid", fgColor=_NAVY)
_HIGHLIGHT_FILL = PatternFill("solid", fgColor="E7EEFB")
_ROW_BORDER = Border(bottom=Side(style="thin", color="E4E9F1"))
_TOTAL_BORDER = Border(top=Side(style="thin", color="B8C0CC"))
_ACCENT_BOTTOM = Border(bottom=Side(style="medium", color=_ACCENT))
_RIGHT = Alignment(horizontal="right", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")

# Series colours for embedded charts (same cycle as the in-app charts).
_CHART_COLORS = ["2F6BD8", "1F9D57", "D98817", "D6453D", "8A93A3",
                 "7C5CFF", "15B8C4", "E6699A"]


def _safe_text(value) -> str:
    """Neutralise CSV/formula injection in a text cell (OWASP guidance).

    Covers the full set of dangerous lead bytes (=, +, -, @, TAB, CR) and looks
    past leading whitespace so " =cmd" cannot slip through. Kept strict so a
    future CSV export inherits the same protection.
    """
    text = "" if value is None else str(value)
    if text.lstrip()[:1] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + text
    return text


def _euros(cents: int):
    return cents_to_euros(int(cents))


# --- Export ----------------------------------------------------------------
def _sheet_title(ws, title: str, subtitle: str, span: int,
                 landscape: bool = False) -> int:
    """Write the navy banner (wordmark look) + subtitle; return the header row.

    Also switches the sheet to a report look: no gridlines, navy tab colour,
    print scaled to one page width (``landscape`` for chart/matrix sheets so a
    print or PDF export never splits columns across pages). The banner spans
    the table width, closed off by an accent baseline — the same signature as
    the app window and the PDF reports.
    """
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _NAVY
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    for col in range(1, span + 1):
        ws.cell(row=1, column=col).fill = _BANNER_FILL
        cell2 = ws.cell(row=2, column=col)
        cell2.fill = _BANNER_FILL
        cell2.border = _ACCENT_BOTTOM
    t = ws.cell(row=1, column=1, value=title)
    t.font = _TITLE_FONT
    t.alignment = Alignment(horizontal="left", vertical="bottom", indent=1)
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = _SUBTITLE_FONT
    s.alignment = Alignment(horizontal="left", vertical="top", indent=1)
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16
    return 4  # row 3 is a spacer; the header goes on row 4


def _header(ws, headers: list[str], row: int, right_cols=(), freeze: bool = True) -> None:
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _RIGHT if (col - 1) in right_cols else _LEFT
    ws.row_dimensions[row].height = 18
    if freeze:
        ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _widths(ws, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _cell(ws, row, col, value, *, money=False, pct=False, zebra=False):
    """Write one body cell with the shared border / alignment / zebra styling."""
    cell = ws.cell(row=row, column=col, value=value)
    if money:
        cell.number_format = _MONEY_FMT
        cell.alignment = _RIGHT
    elif pct:
        cell.number_format = _PCT_FMT
        cell.alignment = _RIGHT
    cell.border = _ROW_BORDER
    if zebra:
        cell.fill = _ALT_FILL
    return cell


def _totals(ws, row: int, span: int, label: str, total_cents: int, money_col: int) -> None:
    for col in range(1, span + 1):
        c = ws.cell(row=row, column=col)
        c.fill = _TOTAL_FILL
        c.border = _TOTAL_BORDER
    ws.cell(row=row, column=1, value=label).font = _TOTAL_FONT
    mc = ws.cell(row=row, column=money_col, value=_euros(total_cents))
    mc.font = _TOTAL_FONT
    mc.number_format = _MONEY_FMT
    mc.alignment = _RIGHT


def _autofilter(ws, header_row: int, last_row: int, span: int) -> None:
    """Clickable column filters over the data block (report convenience)."""
    if last_row > header_row:
        ws.auto_filter.ref = (f"A{header_row}:"
                              f"{get_column_letter(span)}{last_row}")


def _add_pie(ws, *, header_row: int, last_row: int, anchor: str, title: str) -> None:
    """Pie chart over (category, amount) rows, slices in the app's colours."""
    if last_row <= header_row:
        return
    pie = PieChart()
    pie.title = title
    pie.height, pie.width = 9.2, 16.5
    data = Reference(ws, min_col=2, min_row=header_row, max_row=last_row)
    cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=last_row)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats)
    points = []
    for i in range(last_row - header_row):
        dp = DataPoint(idx=i)
        dp.graphicalProperties = GraphicalProperties(
            solidFill=_CHART_COLORS[i % len(_CHART_COLORS)])
        points.append(dp)
    pie.series[0].data_points = points
    ws.add_chart(pie, anchor)


def _add_bars(ws, *, header_row: int, last_row: int, value_cols: list[int],
              anchor: str, title: str, colors: list[str]) -> None:
    """Clustered bar chart: one series per value column, app colour cycle."""
    if last_row <= header_row:
        return
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.height, chart.width = 8.4, 21
    chart.gapWidth = 60
    chart.y_axis.numFmt = '#,##0 "€"'
    # openpyxl quirk: without an explicit delete=False Excel drops the axis
    # labels entirely when a number format is set.
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    for i, col in enumerate(value_cols):
        data = Reference(ws, min_col=col, min_row=header_row, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
        chart.series[i].graphicalProperties = GraphicalProperties(
            solidFill=colors[i % len(colors)])
    chart.set_categories(
        Reference(ws, min_col=1, min_row=header_row + 1, max_row=last_row))
    if len(value_cols) == 1:
        chart.legend = None
    ws.add_chart(chart, anchor)


def export_workbook(
    path: str | Path, *, income, fixed_costs, expenses, credits, overview,
    month_label: str | None = None, history_expenses=None,
) -> Path:
    """Write the multi-sheet report to ``path`` and return it.

    Sheets: Übersicht, Einnahmen, Fixkosten, Fixkosten-Timeline, Ausgaben,
    Ausgaben nach Kategorie, Monatsverlauf, Kredite, Tilgungspläne. Every text
    cell that holds user data goes through ``_safe_text`` (formula-injection
    guard) and all amounts use a euro number format.

    When ``month_label`` is given (e.g. "Juli 2026") the workbook is a monthly
    overview: ``overview`` and ``expenses`` describe that month, so the Übersicht,
    Ausgaben and category sheets are month-specific. ``history_expenses`` (the
    full expense list) still feeds the Monatsverlauf sheet so the multi-month
    trend is preserved; it defaults to ``expenses`` for the old whole-database call.
    """
    hist = history_expenses if history_expenses is not None else expenses
    wb = Workbook()
    today = dates.format_date(dates.today())
    overview_sub = (f"Monat {month_label} · erstellt am {today}"
                    if month_label else f"Erstellt am {today}")

    # --- Übersicht ---
    ws = wb.active
    ws.title = "Übersicht"
    hr = _sheet_title(ws, "HaushaltsManager – Finanzübersicht", overview_sub, 2)
    _widths(ws, [34, 20])
    _header(ws, ["Position", "Betrag / Monat"], hr, right_cols=(1,))
    rows = [
        ("Einnahmen (aktiv)", overview.income_cents, False),
        ("Fixkosten gesamt", overview.fixed_cents, False),
        ("davon Kredite", overview.credits_cents, False),
        ("Variable Ausgaben", overview.variable_cents, False),
        ("Verfügbar nach Fixkosten", overview.after_fixed_cents, True),
        ("Verfügbar nach allem", overview.after_all_cents, True),
    ]
    r = hr + 1
    for i, (label, cents, highlight) in enumerate(rows):
        z = i % 2 == 1
        name = _cell(ws, r, 1, _safe_text(label), zebra=z)
        value = _cell(ws, r, 2, _euros(cents), money=True, zebra=z)
        if highlight:
            # The two "Verfügbar" results are what the report is about —
            # lift them out with a soft accent fill and traffic-light colour.
            name.fill = _HIGHLIGHT_FILL
            name.font = _TOTAL_FONT
            value.fill = _HIGHLIGHT_FILL
            value.font = Font(bold=True, size=12,
                              color=_GREEN if cents >= 0 else _RED)
        r += 1
    # Annual projection uses the RECURRING surplus only — a one-off credit this
    # month must not be multiplied across the whole year (matches the dashboard).
    note = ws.cell(row=r + 1, column=1,
                   value=f"≈ {format_eur(max(0, overview.recurring_after_all_cents) * 12)} pro Jahr ansparbar")
    note.font = _NOTE_FONT

    # --- Einnahmen ---
    ws = wb.create_sheet("Einnahmen")
    hr = _sheet_title(ws, "Einnahmen", "Aktive und inaktive Einnahmequellen", 4)
    _widths(ws, [30, 16, 18, 12])
    _header(ws, ["Bezeichnung", "Art", "Betrag / Monat", "Status"], hr, right_cols=(2,))
    r, total = hr + 1, 0
    for i, it in enumerate(income):
        z = i % 2 == 1
        _cell(ws, r, 1, _safe_text(it.name), zebra=z)
        _cell(ws, r, 2, INCOME_TYPE_LABELS.get(it.income_type, it.income_type), zebra=z)
        _cell(ws, r, 3, _euros(it.amount_cents), money=True, zebra=z)
        _cell(ws, r, 4, "Aktiv" if it.active else "Inaktiv", zebra=z)
        if it.active:
            total += it.amount_cents
        r += 1
    _totals(ws, r, 4, "Summe aktiv", total, 3)
    _autofilter(ws, hr, r - 1, 4)

    # --- Fixkosten ---
    ws = wb.create_sheet("Fixkosten")
    hr = _sheet_title(ws, "Fixkosten", "Laufende monatliche Fixkosten", 6)
    _widths(ws, [28, 16, 16, 14, 14, 16])
    _header(ws, ["Bezeichnung", "Kategorie", "Betrag / Monat", "Beginn", "Ende", "Restlaufzeit"],
            hr, right_cols=(2,))
    r, total = hr + 1, 0
    for i, fc in enumerate(fixed_costs):
        z = i % 2 == 1
        rem = fc.months_remaining()
        _cell(ws, r, 1, _safe_text(fc.name), zebra=z)
        _cell(ws, r, 2, _safe_text(fc.category), zebra=z)
        _cell(ws, r, 3, _euros(fc.amount_cents), money=True, zebra=z)
        _cell(ws, r, 4, dates.format_date(fc.start_date) or "—", zebra=z)
        _cell(ws, r, 5, dates.format_date(fc.end_date) or "unbegrenzt", zebra=z)
        _cell(ws, r, 6, dates.format_months_remaining(rem) if rem is not None else "unbegrenzt", zebra=z)
        total += fc.amount_cents
        r += 1
    _totals(ws, r, 6, "Summe", total, 3)
    _autofilter(ws, hr, r - 1, 6)

    # --- Fixkosten-Timeline ---
    ws = wb.create_sheet("Fixkosten-Timeline")
    hr = _sheet_title(ws, "Fixkosten-Abbau", "Wann Fixkosten wegfallen und was übrig bleibt", 5)
    _widths(ws, [16, 34, 16, 16, 18])
    _header(ws, ["Datum", "Wegfallende Kosten", "Betrag", "Neue Fixsumme", "Verfügbar danach"],
            hr, right_cols=(2, 3, 4))
    result = timeline.build(overview.recurring_income_cents, list(fixed_costs))
    r = hr + 1
    _cell(ws, r, 1, "Jetzt")
    _cell(ws, r, 2, "—")
    _cell(ws, r, 3, _euros(0), money=True)
    _cell(ws, r, 4, _euros(result.current_fixed_total_cents), money=True)
    _cell(ws, r, 5, _euros(result.current_available_cents), money=True)
    r += 1
    for i, ev in enumerate(result.events):
        z = i % 2 == 0
        _cell(ws, r, 1, ev.label, zebra=z)
        _cell(ws, r, 2, _safe_text(", ".join(d.name for d in ev.dropped)), zebra=z)
        _cell(ws, r, 3, _euros(ev.dropped_amount_cents), money=True, zebra=z)
        _cell(ws, r, 4, _euros(ev.new_fixed_total_cents), money=True, zebra=z)
        _cell(ws, r, 5, _euros(ev.available_after_fixed_cents), money=True, zebra=z)
        r += 1

    # --- Ausgaben ---
    ws = wb.create_sheet("Ausgaben")
    exp_sub = f"Erfasste Ausgaben im Monat {month_label}" if month_label else "Einzelne erfasste Ausgaben"
    hr = _sheet_title(ws, "Variable Ausgaben", exp_sub, 4)
    _widths(ws, [16, 26, 42, 16])
    _header(ws, ["Datum", "Kategorie", "Beschreibung", "Betrag"], hr, right_cols=(3,))
    r, total = hr + 1, 0
    for i, e in enumerate(expenses):
        z = i % 2 == 1
        desc = e.description or ""
        if getattr(e, "recurring", False):
            desc = (f"{desc}  (monatlich)").strip()
        _cell(ws, r, 1, dates.format_date(e.date), zebra=z)
        _cell(ws, r, 2, _safe_text(e.category), zebra=z)
        _cell(ws, r, 3, _safe_text(desc), zebra=z)
        _cell(ws, r, 4, _euros(e.amount_cents), money=True, zebra=z)
        total += e.amount_cents
        r += 1
    _totals(ws, r, 4, "Summe", total, 4)
    _autofilter(ws, hr, r - 1, 4)

    # --- Ausgaben nach Kategorie ---
    by_cat: dict[str, int] = defaultdict(int)
    for e in expenses:
        by_cat[e.category] += e.amount_cents
    ws = wb.create_sheet("Ausgaben nach Kategorie")
    cat_sub = (f"Summe je Kategorie im Monat {month_label}"
               if month_label else "Summe je Kategorie über alle erfassten Ausgaben")
    hr = _sheet_title(ws, "Ausgaben nach Kategorie", cat_sub, 3, landscape=True)
    _widths(ws, [24, 18, 12])
    _header(ws, ["Kategorie", "Betrag", "Anteil"], hr, right_cols=(1, 2))
    cat_total = sum(by_cat.values()) or 1
    r = hr + 1
    for i, (cat, cents) in enumerate(sorted(by_cat.items(), key=lambda kv: kv[1], reverse=True)):
        z = i % 2 == 1
        _cell(ws, r, 1, _safe_text(cat), zebra=z)
        _cell(ws, r, 2, _euros(cents), money=True, zebra=z)
        _cell(ws, r, 3, cents / cat_total, pct=True, zebra=z)
        r += 1
    _totals(ws, r, 3, "Summe", sum(by_cat.values()), 2)
    _add_pie(ws, header_row=hr, last_row=r - 1, anchor="E4",
             title="Anteile" + (f" im Monat {month_label}" if month_label else ""))

    # --- Monatsverlauf (always the full history, independent of a chosen month) ---
    monthly: dict[str, int] = defaultdict(int)
    for e in hist:
        monthly[(e.date or "")[:7]] += e.amount_cents
    ws = wb.create_sheet("Monatsverlauf")
    hr = _sheet_title(ws, "Monatsverlauf", "Variable Ausgaben je Monat", 2, landscape=True)
    _widths(ws, [18, 20])
    _header(ws, ["Monat", "Variable Ausgaben"], hr, right_cols=(1,))
    r = hr + 1
    for i, ym in enumerate(sorted(k for k in monthly if k)):
        z = i % 2 == 1
        label = f"{ym[5:7]}/{ym[0:4]}" if len(ym) >= 7 else ym
        _cell(ws, r, 1, label, zebra=z)
        _cell(ws, r, 2, _euros(monthly[ym]), money=True, zebra=z)
        r += 1
    _add_bars(ws, header_row=hr, last_row=r - 1, value_cols=[2], anchor="D4",
              title="Variable Ausgaben je Monat", colors=[_ACCENT])

    # --- Kredite ---
    ws = wb.create_sheet("Kredite")
    hr = _sheet_title(ws, "Kredite", "Laufende und abbezahlte Kredite", 7, landscape=True)
    _widths(ws, [26, 14, 16, 14, 11, 14, 12])
    _header(ws, ["Bezeichnung", "Kategorie", "Gesamtbetrag", "Rate / Monat", "Zins p.a.", "Ende", "Status"],
            hr, right_cols=(2, 3, 4))
    r, monthly_active = hr + 1, 0
    for i, cr in enumerate(credits):
        z = i % 2 == 1
        _cell(ws, r, 1, _safe_text(cr.name), zebra=z)
        _cell(ws, r, 2, _safe_text(cr.category), zebra=z)
        _cell(ws, r, 3, _euros(cr.total_cents or 0), money=True, zebra=z)
        _cell(ws, r, 4, _euros(cr.monthly_cents or 0), money=True, zebra=z)
        zins = _cell(ws, r, 5, (cr.interest_rate / 100.0) if cr.interest_rate else None, zebra=z)
        zins.number_format = '0.00 %'
        zins.alignment = _RIGHT
        _cell(ws, r, 6, dates.format_date(cr.end_date) or "—", zebra=z)
        _cell(ws, r, 7, CREDIT_STATUS_LABELS.get(cr.status, cr.status), zebra=z)
        if cr.status == "aktiv" and cr.monthly_cents:
            monthly_active += cr.monthly_cents
        r += 1
    _totals(ws, r, 7, "Monatliche Belastung (aktiv)", monthly_active, 4)
    _autofilter(ws, hr, r - 1, 7)

    # --- Tilgungspläne ---
    ws = wb.create_sheet("Tilgungspläne")
    _write_amortisation_plans(ws, credits)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return Path(path)


def export_year_workbook(path: str | Path, *, overview, prev_overview=None) -> Path:
    """Write the annual report (Jahresbericht) to ``path`` and return it.

    Sheets: Jahresübersicht (totals + Sparquote, optionally with the previous
    year and the delta), Monate (per-month income/expenses/balance) and
    Monatsmatrix (category × month). ``overview``/``prev_overview`` are
    :class:`modules.annual.YearOverview` instances; ``prev_overview`` may be
    None or empty, then the comparison columns are omitted.
    """
    from modules import annual as _annual  # local import to avoid cycle

    wb = Workbook()
    today = dates.format_date(dates.today())
    year = overview.year
    has_prev = prev_overview is not None and prev_overview.has_data
    covered_note = (f"Januar bis {dates.month_name(overview.months_covered)}"
                    if overview.months_covered < 12 else "ganzes Jahr")

    # --- Jahresübersicht ---
    ws = wb.active
    ws.title = "Jahresübersicht"
    span = 4 if has_prev else 2
    hr = _sheet_title(ws, f"HaushaltsManager – Jahresbericht {year}",
                      f"{covered_note} · erstellt am {today}", span)
    _widths(ws, [30, 18, 18, 18])
    headers = ["Position", str(year)] + ([str(prev_overview.year), "Δ zum Vorjahr"] if has_prev else [])
    _header(ws, headers, hr, right_cols=(1, 2, 3))
    rows = [
        ("Einnahmen", overview.income_cents,
         prev_overview.income_cents if has_prev else None),
        ("Fixkosten", overview.fixed_cents,
         prev_overview.fixed_cents if has_prev else None),
        ("Variable Ausgaben", overview.variable_cents,
         prev_overview.variable_cents if has_prev else None),
        ("Ausgaben gesamt", overview.expenses_cents,
         prev_overview.expenses_cents if has_prev else None),
        ("Saldo (gespart)", overview.remaining_cents,
         prev_overview.remaining_cents if has_prev else None),
    ]
    r = hr + 1
    for i, (label, cur, prev) in enumerate(rows):
        z = i % 2 == 1
        cells = [_cell(ws, r, 1, _safe_text(label), zebra=z),
                 _cell(ws, r, 2, _euros(cur), money=True, zebra=z)]
        if has_prev:
            cells.append(_cell(ws, r, 3, _euros(prev), money=True, zebra=z))
            cells.append(_cell(ws, r, 4, _euros(cur - prev), money=True, zebra=z))
        if label.startswith("Saldo"):
            # The year's bottom line — lifted out like on the dashboard.
            for cell in cells:
                cell.fill = _HIGHLIGHT_FILL
                cell.font = _TOTAL_FONT
            cells[1].font = Font(bold=True, size=12,
                                 color=_GREEN if cur >= 0 else _RED)
        r += 1
    # Sparquote as a percent row (rate of a zero income stays empty).
    z = len(rows) % 2 == 1
    _cell(ws, r, 1, "Sparquote", zebra=z)
    _cell(ws, r, 2, overview.savings_rate, pct=True, zebra=z)
    if has_prev:
        _cell(ws, r, 3, prev_overview.savings_rate, pct=True, zebra=z)
        if overview.savings_rate is not None and prev_overview.savings_rate is not None:
            _cell(ws, r, 4, overview.savings_rate - prev_overview.savings_rate,
                  pct=True, zebra=z)

    # --- Monate ---
    ws = wb.create_sheet("Monate")
    hr = _sheet_title(ws, f"Monate {year}", "Einnahmen, Ausgaben und Saldo je Monat", 5,
                      landscape=True)
    _widths(ws, [16, 16, 16, 18, 16])
    _header(ws, ["Monat", "Einnahmen", "Fixkosten", "Variable Ausgaben", "Saldo"],
            hr, right_cols=(1, 2, 3, 4))
    r = hr + 1
    for i, pt in enumerate(overview.points):
        z = i % 2 == 1
        _cell(ws, r, 1, dates.month_name(pt.month), zebra=z)
        _cell(ws, r, 2, _euros(pt.income_cents), money=True, zebra=z)
        _cell(ws, r, 3, _euros(pt.fixed_cents), money=True, zebra=z)
        _cell(ws, r, 4, _euros(pt.variable_cents), money=True, zebra=z)
        _cell(ws, r, 5, _euros(pt.remaining_cents), money=True, zebra=z)
        r += 1
    _totals(ws, r, 5, "Summe Saldo", overview.remaining_cents, 5)
    _add_bars(ws, header_row=hr, last_row=r - 1, value_cols=[2, 3, 4], anchor="G4",
              title=f"Einnahmen und Ausgaben {year}",
              colors=[_ACCENT, "D98817", "8A93A3"])

    # --- Monatsmatrix (Kategorie × Monat) ---
    categories, per_month = _annual.category_matrix(overview)
    ws = wb.create_sheet("Monatsmatrix")
    hr = _sheet_title(ws, f"Ausgaben-Matrix {year}",
                      "Variable Ausgaben je Kategorie und Monat", 14,
                      landscape=True)
    _widths(ws, [24] + [11] * 12 + [13])
    month_heads = [dates.month_name(m)[:3] for m in range(1, 13)]
    _header(ws, ["Kategorie"] + month_heads + ["Summe"], hr,
            right_cols=tuple(range(1, 14)))
    r = hr + 1
    for i, cat in enumerate(categories):
        z = i % 2 == 1
        _cell(ws, r, 1, _safe_text(cat), zebra=z)
        for m in range(12):
            _cell(ws, r, 2 + m, _euros(per_month[cat][m]), money=True, zebra=z)
        _cell(ws, r, 14, _euros(sum(per_month[cat])), money=True, zebra=z)
        r += 1
    _totals(ws, r, 14, "Summe", overview.variable_cents, 14)

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return Path(path)


def _write_amortisation_plans(ws, credits) -> None:
    """Best-effort amortisation tables for credits that have enough figures."""
    hr = _sheet_title(ws, "Tilgungspläne",
                      "Berechnete Tilgung je Kredit mit ausreichenden Angaben", 5)
    _widths(ws, [10, 16, 16, 16, 18])
    row = hr
    any_plan = False
    for cr in credits:
        principal = cr.total_cents
        term = cr.term_months
        if not term and cr.end_date and cr.start_date:
            sd, ed = dates.parse_date(cr.start_date), dates.parse_date(cr.end_date)
            if sd and ed:
                term = max(1, dates.months_between(sd, ed))
        if not principal and cr.monthly_cents and term:
            principal = cr.monthly_cents * term  # 0 %-Annahme
        if not (principal and term):
            continue

        any_plan = True
        rate = cr.interest_rate or 0.0
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=_safe_text(cr.name)).font = _TOTAL_FONT
        row += 1
        _header(ws, ["Monat", "Rate", "Tilgung", "Zinsen", "Restschuld"], row,
                right_cols=(1, 2, 3, 4), freeze=False)
        row += 1
        for j, sr in enumerate(annuity.build_schedule(principal, rate, term)):
            z = j % 2 == 1
            _cell(ws, row, 1, sr.month, zebra=z)
            _cell(ws, row, 2, _euros(sr.payment_cents), money=True, zebra=z)
            _cell(ws, row, 3, _euros(sr.principal_cents), money=True, zebra=z)
            _cell(ws, row, 4, _euros(sr.interest_cents), money=True, zebra=z)
            _cell(ws, row, 5, _euros(sr.balance_cents), money=True, zebra=z)
            row += 1
        row += 1  # blank spacer between plans
    if not any_plan:
        ws.cell(row=hr, column=1,
                value="Keine Kredite mit ausreichenden Angaben für einen Tilgungsplan.")


# --- Import ----------------------------------------------------------------
# Header keywords -> logical field. Lower-cased substring match.
_COLUMN_HINTS = {
    "date": ["datum", "buchungstag", "valuta", "wertstellung", "buchung"],
    "amount": ["betrag", "umsatz", "wert", "soll", "haben", "amount"],
    "description": ["verwendungszweck", "beschreibung", "buchungstext", "umsatztext",
                    "vorgang", "name", "empfänger", "beguenstigter", "auftraggeber"],
    "category": ["kategorie", "category"],
}


def guess_mapping(headers: list[str]) -> dict[str, int | None]:
    """Guess which column index holds date/amount/description/category."""
    mapping: dict[str, int | None] = {"date": None, "amount": None,
                                      "description": None, "category": None}
    lowered = [(i, str(h).strip().lower()) for i, h in enumerate(headers)]
    for field, hints in _COLUMN_HINTS.items():
        for i, head in lowered:
            if any(hint in head for hint in hints):
                mapping[field] = i
                break
    return mapping


def read_preview(path: str | Path, max_rows: int = 500) -> tuple[list[str], list[list], dict]:
    """Read the first sheet: return (headers, data_rows, guessed_mapping)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], [], {"date": None, "amount": None, "description": None, "category": None}
    headers = [("" if h is None else str(h)) for h in rows[0]]
    data = [list(r) for r in rows[1: 1 + max_rows]]
    return headers, data, guess_mapping(headers)


def rows_to_expenses(data_rows: list[list], mapping: dict, default_category: str = "Sonstiges"):
    """Convert mapped rows into VariableExpense models, skipping unparseable ones.

    Returns (expenses, skipped_count). Amounts are taken as absolute values so a
    bank export's negative debits import cleanly as expenses.
    """
    from modules.models import VariableExpense  # local import to avoid cycle

    out = []
    skipped = 0
    di, ai = mapping.get("date"), mapping.get("amount")
    desc_i, cat_i = mapping.get("description"), mapping.get("category")
    for row in data_rows:
        try:
            if ai is None or ai >= len(row):
                skipped += 1
                continue
            cents = parse_eur(row[ai])
            if cents == 0:
                skipped += 1
                continue
            d = dates.parse_date(row[di]) if di is not None and di < len(row) else None
            iso = dates.to_iso(d) if d else dates.to_iso(dates.today())
            desc = str(row[desc_i]) if desc_i is not None and desc_i < len(row) and row[desc_i] else ""
            cat = str(row[cat_i]) if cat_i is not None and cat_i < len(row) and row[cat_i] else default_category
            out.append(VariableExpense(
                date=iso, amount_cents=abs(cents), category=cat or default_category,
                description=desc.strip()))
        except Exception:  # noqa: BLE001 - a bad row must not abort the import
            skipped += 1
    return out, skipped
